import openpyxl

list=[]
dict = {}
# to load workbook
book=openpyxl.load_workbook("E:\\Learnings\\Python\\SeleniumpythonFramework\\TestData.xlsx")

# to select the active sheet
sheet=book.active

# to read a particular cell
cell=sheet.cell(row=2,column=1)

# to write a particular cell
#sheet.cell(row=2,column=3).value = "Test Lead"


# to loop each row and col
for i in range(2,sheet.max_row+1):
    dict.clear()
    for j in range(1, sheet.max_column + 1):
        dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
        print(dict)
    list.append(dict)
    #print [dict]

print(list)